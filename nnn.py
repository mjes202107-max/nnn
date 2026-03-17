#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
課表生成程式
生成課表並存為 Excel 檔案
"""

import csv
import zipfile
import xml.etree.ElementTree as ET
from datetime import datetime
import os

# 嘗試導入 openpyxl
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# 課表資料
schedule_data = [
    {"time": "09:00~09:30", "content": "報到", "speaker": "無"},
    {"time": "09:30~09:50", "content": "開場致詞", "speaker": "林院長"},
    {"time": "09:50~10:20", "content": "5G 專網與企業數位轉型的實踐之路", "speaker": "張教授"},
    {"time": "10:20~10:50", "content": "AI 驅動的智慧工廠無線網路架構設計", "speaker": "李教授"},
    {"time": "10:50~11:10", "content": "Break", "speaker": "無"},
    {"time": "11:10~11:40", "content": "Open RAN 生態系統的發展與挑戰", "speaker": "研究團隊"},
    {"time": "11:40~12:10", "content": "雲原生架構在 5G 核心網路中的應用", "speaker": "研究團隊"},
    {"time": "12:10~13:30", "content": "Lunch", "speaker": "無"},
    {"time": "13:30~14:00", "content": "毫米波通訊技術與室內定位整合應用", "speaker": "研究團隊"},
    {"time": "14:00~14:30", "content": "網路切片技術於醫療場域的實證研究", "speaker": "胡教授"},
    {"time": "14:30~14:50", "content": "Break", "speaker": "無"},
    {"time": "14:50~15:30", "content": "O-RAN 近即時控制器 xApp 開發實戰", "speaker": "研究團隊"},
    {"time": "15:30~16:00", "content": "綜合座談與交流", "speaker": "研究團隊"},
    {"time": "16:00~16:20", "content": "Break", "speaker": "無"},
    {"time": "16:20~17:00", "content": "Energy-Efficient Resource Allocation in O-RAN Architecture", "speaker": "賴博士"},
    {"time": "17:00~17:40", "content": "Deep Reinforcement Learning for Network Slicing Optimization", "speaker": "賴博士"},
    {"time": "17:40~18:10", "content": "Dinner", "speaker": "無"},
    {"time": "18:10~18:50", "content": "Federated Learning Approaches for Privacy-Preserving 6G Networks", "speaker": "陳教授"},
    {"time": "18:50~19:30", "content": "Digital Twin-Enabled Intelligent RAN Management", "speaker": "陳教授"},
    {"time": "19:30~20:00", "content": "Panel Discussion and Closing Remarks", "speaker": "研究團隊"},
]


def create_schedule_excel(filename="nnn.xlsx"):
    """
    建立課表 Excel 檔案
    
    Args:
        filename (str): 輸出檔案名稱
    """
    if HAS_OPENPYXL:
        # 使用 openpyxl 建立格式化的 Excel 檔案
        wb = Workbook()
        ws = wb.active
        ws.title = "課表"
        
        # 設定欄位寬度
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 15
        
        # 設定標題列
        headers = ["時間", "內容", "講者"]
        
        # 標題樣式
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        center_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 寫入標題
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = thin_border
        
        # 寫入數據
        for row_idx, item in enumerate(schedule_data, start=2):
            # 時間
            time_cell = ws.cell(row=row_idx, column=1)
            time_cell.value = item["time"]
            time_cell.alignment = center_alignment
            time_cell.border = thin_border
            
            # 內容
            content_cell = ws.cell(row=row_idx, column=2)
            content_cell.value = item["content"]
            content_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            content_cell.border = thin_border
            
            # 講者
            speaker_cell = ws.cell(row=row_idx, column=3)
            speaker_cell.value = item["speaker"]
            speaker_cell.alignment = center_alignment
            speaker_cell.border = thin_border
        
        # 設定列高
        ws.row_dimensions[1].height = 25
        for row_idx in range(2, len(schedule_data) + 2):
            ws.row_dimensions[row_idx].height = 30
        
        # 保存檔案
        wb.save(filename)
        print(f"課表已保存為 {filename}")
    else:
        # 手動建立 XLSX 檔案
        create_xlsx_manual(filename)


def create_xlsx_manual(filename="nnn.xlsx"):
    """
    手動建立 XLSX 檔案（當 openpyxl 不可用時）
    """
    # 建立臨時目錄結構
    import tempfile
    import shutil
    
    temp_dir = tempfile.mkdtemp()
    
    try:
        # 建立目錄結構
        os.makedirs(os.path.join(temp_dir, "_rels"), exist_ok=True)
        os.makedirs(os.path.join(temp_dir, "xl", "worksheets"), exist_ok=True)
        os.makedirs(os.path.join(temp_dir, "xl", "_rels"), exist_ok=True)
        os.makedirs(os.path.join(temp_dir, "docProps"), exist_ok=True)
        
        # 1. 建立 [Content_Types].xml
        content_types = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
    <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
    <Default Extension="xml" ContentType="application/xml"/>
    <Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>
    <Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>
    <Override PartName="/xl/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/>
    <Override PartName="/docProps/core.xml" ContentType="application/vnd.openxmlformats-package.core-properties+xml"/>
</Types>'''
        
        with open(os.path.join(temp_dir, "[Content_Types].xml"), 'w', encoding='utf-8') as f:
            f.write(content_types)
        
        # 2. 建立 _rels/.rels
        rels = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
    <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>
    <Relationship Id="rId2" Type="http://schemas.openxmlformats.org/package/2006/relationships/metadata/core-properties" Target="docProps/core.xml"/>
</Relationships>'''
        
        with open(os.path.join(temp_dir, "_rels", ".rels"), 'w', encoding='utf-8') as f:
            f.write(rels)
        
        # 3. 建立 xl/_rels/workbook.xml.rels
        workbook_rels = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
    <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>
    <Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/>
</Relationships>'''
        
        with open(os.path.join(temp_dir, "xl", "_rels", "workbook.xml.rels"), 'w', encoding='utf-8') as f:
            f.write(workbook_rels)
        
        # 4. 建立 xl/styles.xml
        styles_xml = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
    <fonts count="2">
        <font><sz val="11"/><color theme="1"/><name val="Calibri"/><family val="2"/></font>
        <font><b/><sz val="11"/><color rgb="FFFFFFFF"/><name val="Calibri"/><family val="2"/></font>
    </fonts>
    <fills count="3">
        <fill><patternFill patternType="none"/></fill>
        <fill><patternFill patternType="gray125"/></fill>
        <fill><patternFill patternType="solid"><fgColor rgb="FF4472C4"/></patternFill></fill>
    </fills>
    <borders count="2">
        <border><left/><right/><top/><bottom/><diagonal/></border>
        <border><left style="thin"/><right style="thin"/><top style="thin"/><bottom style="thin"/><diagonal/></border>
    </borders>
    <cellStyleXfs count="1">
        <xf numFmtId="0" fontId="0" fillId="0" borderId="0"/>
    </cellStyleXfs>
    <cellXfs count="4">
        <xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/>
        <xf numFmtId="0" fontId="1" fillId="2" borderId="1" xfId="0" applyFill="1" applyFont="1" applyBorder="1"/>
        <xf numFmtId="0" fontId="0" fillId="0" borderId="1" xfId="0" applyBorder="1"/>
        <xf numFmtId="0" fontId="0" fillId="0" borderId="1" xfId="0" alignment="horizontal" applyBorder="1"/>
    </cellXfs>
</styleSheet>'''
        
        with open(os.path.join(temp_dir, "xl", "styles.xml"), 'w', encoding='utf-8') as f:
            f.write(styles_xml)
        
        # 5. 建立 xl/workbook.xml
        workbook_xml = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
    <sheets>
        <sheet name="課表" sheetId="1" r:id="rId1"/>
    </sheets>
</workbook>'''
        
        with open(os.path.join(temp_dir, "xl", "workbook.xml"), 'w', encoding='utf-8') as f:
            f.write(workbook_xml)
        
        # 6. 建立 xl/worksheets/sheet1.xml
        cells_xml = '        <sheetData>\n'
        cells_xml += '            <row r="1" spans="1:3" ht="25">\n'
        headers = ["時間", "內容", "講者"]
        for col_idx, header in enumerate(headers, start=1):
            cell_ref = chr(64 + col_idx) + "1"
            cells_xml += f'                <c r="{cell_ref}" s="1" t="inlineStr"><is><t>{header}</t></is></c>\n'
        cells_xml += '            </row>\n'
        
        for row_idx, item in enumerate(schedule_data, start=2):
            row_height = 30
            cells_xml += f'            <row r="{row_idx}" spans="1:3" ht="{row_height}">\n'
            
            # 時間
            cell_ref = "A" + str(row_idx)
            cells_xml += f'                <c r="{cell_ref}" s="2" t="inlineStr"><is><t>{item["time"]}</t></is></c>\n'
            
            # 內容
            cell_ref = "B" + str(row_idx)
            cells_xml += f'                <c r="{cell_ref}" s="2" t="inlineStr"><is><t>{item["content"]}</t></is></c>\n'
            
            # 講者
            cell_ref = "C" + str(row_idx)
            cells_xml += f'                <c r="{cell_ref}" s="2" t="inlineStr"><is><t>{item["speaker"]}</t></is></c>\n'
            
            cells_xml += '            </row>\n'
        
        cells_xml += '        </sheetData>\n'
        
        sheet_xml = f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
    <sheetPr><outlinePr summaryBelow="1" summaryRight="1"/></sheetPr>
    <sheetViews><sheetView tabSelected="1" workbookViewId="0"/></sheetViews>
    <sheetFormatPr defaultRowHeight="15"/>
    <cols>
        <col min="1" max="1" width="15" bestFit="1"/>
        <col min="2" max="2" width="40" bestFit="1"/>
        <col min="3" max="3" width="15" bestFit="1"/>
    </cols>
{cells_xml}
</worksheet>'''
        
        with open(os.path.join(temp_dir, "xl", "worksheets", "sheet1.xml"), 'w', encoding='utf-8') as f:
            f.write(sheet_xml)
        
        # 7. 建立 docProps/core.xml
        core_props = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<cp:coreProperties xmlns:cp="http://schemas.openxmlformats.org/officeDocument/2006/metadata/core-properties" xmlns:dc="http://purl.org/dc/elements/1.1/" xmlns:dcterms="http://purl.org/dc/terms/" xmlns:dcmitype="http://purl.org/dc/dcmitype/" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">
    <dc:title>課表</dc:title>
    <dc:creator>課表生成程式</dc:creator>
    <cp:lastModifiedBy>課表生成程式</cp:lastModifiedBy>
    <dcterms:created xsi:type="dcterms:W3CDTF">2026-03-17T00:00:00Z</dcterms:created>
    <dcterms:modified xsi:type="dcterms:W3CDTF">2026-03-17T00:00:00Z</dcterms:modified>
</cp:coreProperties>'''
        
        with open(os.path.join(temp_dir, "docProps", "core.xml"), 'w', encoding='utf-8') as f:
            f.write(core_props)
        
        # 8. 建立 ZIP 檔案（XLSX）
        with zipfile.ZipFile(filename, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for root, dirs, files in os.walk(temp_dir):
                for file in files:
                    file_path = os.path.join(root, file)
                    arcname = os.path.relpath(file_path, temp_dir)
                    zipf.write(file_path, arcname)
        
        print(f"課表已保存為 {filename}")
        
    finally:
        # 清理臨時目錄
        shutil.rmtree(temp_dir)


def print_schedule():
    """
    列印課表到控制台
    """
    print("\n" + "="*70)
    print("課表".center(70))
    print("="*70)
    print(f"{'時間':<15} {'內容':<40} {'講者':<10}")
    print("-"*70)
    
    for item in schedule_data:
        print(f"{item['time']:<15} {item['content']:<40} {item['speaker']:<10}")
    
    print("="*70 + "\n")


if __name__ == "__main__":
    # 列印課表
    print_schedule()
    
    # 生成 Excel 檔案
    create_schedule_excel("nnn.xlsx")
    
    print(f"課表資料統計：總共 {len(schedule_data)} 個時段")
